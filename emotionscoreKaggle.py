import os, sys
from shutil import copy
sys.stdout = open('Output.txt', 'w')



def detect_faces(path):
    """Detects faces in an image."""
    from google.cloud import vision
    import io, os
    
    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = "../input/apikey/Key.json"

    client = vision.ImageAnnotatorClient()

    with io.open(path, 'rb') as image_file:
        content = image_file.read()

    image = vision.Image(content=content)

    response = client.face_detection(image=image)
    faces = response.face_annotations

    # Names of likelihood from google.cloud.vision.enums
    likelihood_name = ('UNKNOWN', 'VERY_UNLIKELY', 'UNLIKELY', 'POSSIBLE', 'LIKELY', 'VERY_LIKELY')
    flag = 0
    imageBool = 0
    # Iterating over faces in image (will be executed only once as there should only be one face in each image)
    for face in faces:
        flag = 1
        d = face.detection_confidence
        a = likelihood_name[face.anger_likelihood]
        j = likelihood_name[face.joy_likelihood]
        su = likelihood_name[face.surprise_likelihood]
        so = likelihood_name[face.sorrow_likelihood]
        h = likelihood_name[face.headwear_likelihood]
        b = likelihood_name[face.blurred_likelihood]
        u = likelihood_name[face.under_exposed_likelihood]
        t = face.tilt_angle
        p = face.pan_angle
        r = face.roll_angle

        # Exporting values to text file
        print('Detection confidence:{}'.format(d))
        print('anger: {}'.format(a))
        print('joy: {}'.format(j))
        print('surprise: {}'.format(su))
        print('sorrow: {}'.format(so))
        print('headwear: {}'.format(h))
        print('blurred: {}'.format(b))
        print('under exposed: {}'.format(u))
        print('tilt angle: {}'.format(t))
        print('pan angle: {}'.format(p))
        print('roll angle: {}'.format(r))

        # Checking if the image meets the critira
        if d > 0.5 and a == 'VERY_UNLIKELY' and j == 'VERY_UNLIKELY' and su == 'VERY_UNLIKELY' and so == 'VERY_UNLIKELY' and h == 'VERY_UNLIKELY' and b == 'VERY_UNLIKELY' and u == 'VERY_UNLIKELY' and  t < 5 and t > -5 and p < 5 and p > -5 and r < 5 and r > -5:
            imageBool = 1
        else:
            imageBool = 0
        
        # Face bounds are not needed, but just in case ¯\_(ツ)_/¯
        vertices = (['({},{})'.format(vertex.x, vertex.y)
                    for vertex in face.bounding_poly.vertices])
        print('face bounds: {} \n'.format(','.join(vertices)))
        break

    # A flag to mark finding at least one face; because some images are so distorted that no face can be detected
    if flag == 0:
        # Appending zero for distorted image; in order not to mess the list ordering
        imageBool = 0
        print('Detection confidence: 0')
        print('anger: UNKNOWN')
        print('joy: UNKNOWN')
        print('surprise: UNKNOWN')
        print('sorrow: UNKNOWN')
        print('headwear: UNKNOWN')
        print('blurred: UNKNOWN')
        print('under exposed: UNKNOWN')
        print('tilt angle: 99')
        print('pan angle: 99')
        print('roll angle: 99')
        
    if imageBool == 1:
        copy(path, './')
        
    # Error Handling
    if response.error.message:
        raise Exception(
            '{}\nFor more info on error messages, check: '
            'https://cloud.google.com/apis/design/errors'.format(
                response.error.message))





!python3 -m pip install openpyxl





from openpyxl import load_workbook

# Change to the path of the directory containing the excel file
wb = load_workbook("../input/testexcel/Output.xlsx")
ws = wb['Sheet1']
columnA = ws['A']
# Getting the paths of the images
done = [columnA[x].value for x in range(len(columnA))]
    
i = 0
flag = 0
for dirname, _, filenames in os.walk('../input/human-faces-dataset/r1/r1'):
    if flag == 1:
        sys.stdout.close()
        break
    for filename in filenames:
        i += 1
        if i == 7802:
            flag = 1
            break
        image = os.path.join(dirname, filename)
        if image not in done:
            print(image)
            detect_faces(image)




import xlsxwriter, itertools 
d = []
a = []
j = []
su = []
so = []
h = []
b = []
u = []
t = []
p = []
r = []
leftImages = []
fullPath = [] 

with open('./Output.txt', 'r') as searchfile:
    for line in searchfile:
        if '../input/human-faces-dataset/' in line:
            fullPath.append(line.strip())
        elif 'Detection confidence:' in line:
            d.append(float(line[21:].strip()))
        elif 'anger: ' in line:
            a.append(line[7:].strip())
        elif 'joy: ' in line:
            j.append(line[5:].strip())
        elif 'surprise: ' in line:
            su.append(line[9:].strip())
        elif 'sorrow: ' in line:
            so.append(line[8:].strip())
        elif 'headwear: ' in line:
            h.append(line[10:].strip())
        elif 'blurred: ' in line:
            b.append(line[9:].strip())
        elif 'under exposed: ' in line:
            u.append(line[15:].strip())
        elif 'tilt angle: ' in line:
            t.append(float(line[12:].strip()))
        elif 'pan angle: ' in line:
            p.append(float(line[11:].strip()))
        elif 'roll angle: ' in line:
            r.append(float(line[12:].strip()))
            
            
outWorkbook = xlsxwriter.Workbook('../input/exceloutput/Output.xlsx')
outSheet = outWorkbook.get_worksheet_by_name('Sheet1')

# Initialize i with the first empty row
i = len(done)
for i in range(len(fullPath)):
    outSheet.write(i + 1, 0, fullPath[i])
    outSheet.write(i + 1, 2, d[i])
    outSheet.write(i + 1, 3, t[i])
    outSheet.write(i + 1, 4, p[i])
    outSheet.write(i + 1, 5, r[i])

for (d, a, j, su, so, h, b, u, t, p, r) in itertools.zip_longest(d, a, j, su, so, h, b, u, t, p, r):
    if d > 0.5 and a == 'VERY_UNLIKELY' and j == 'VERY_UNLIKELY' and su == 'VERY_UNLIKELY' and so == 'VERY_UNLIKELY' and h == 'VERY_UNLIKELY' and b == 'VERY_UNLIKELY' and u == 'VERY_UNLIKELY' and  t < 5 and t > -5 and p < 5 and p > -5 and r < 5 and r > -5:
        leftImages.append('1')
    else:
        leftImages.append('0')
j = 0
for j in range(len(fullPath)):
    outSheet.write(j + 1, 1, leftImages[j])
outWorkbook.close()




#os.remove("/kaggle/working/Output.txt")
