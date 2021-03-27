# pylint: disable=no-member

def detect_faces(uri):
    """Detects faces in an image."""
    from google.cloud import vision

    os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = "/home/ahmad/Downloads/GraduationProject/API/Key.json"

    client = vision.ImageAnnotatorClient()

    image = vision.Image()
    image.source.image_uri = uri

    response = client.face_detection(image=image)
    faces = response.face_annotations

    # Names of likelihood from google.cloud.vision.enums
    likelihood_name = ('UNKNOWN', 'VERY_UNLIKELY', 'UNLIKELY', 'POSSIBLE', 'LIKELY', 'VERY_LIKELY')
    flag = 0
    # Iterating over faces in image (will be executed only once as there should only be one face in each image)
    for face in faces:
        flag = 1
        d = face.detection_confidence
        a = likelihood_name[face.anger_likelihood]
        j = likelihood_name[face.joy_likelihood]
        su = likelihood_name[face.surprise_likelihood]
        so = likelihood_name[face.sorrow_likelihood]
        h = likelihood_name[face.headwear_likelihood]
        b = likelihood_name[face.blurred_likelihood]
        u = likelihood_name[face.under_exposed_likelihood]
        t = face.tilt_angle
        p = face.pan_angle
        r = face.roll_angle

        # Exporting values to text file
        print('Detection confidence:{}'.format(d))
        print('anger: {}'.format(a))
        print('joy: {}'.format(j))
        print('surprise: {}'.format(su))
        print('sorrow: {}'.format(so))
        print('headwear: {}'.format(h))
        print('blurred: {}'.format(b))
        print('under exposed: {}'.format(u))
        print('tilt angle: {}'.format(t))
        print('pan angle: {}'.format(p))
        print('roll angle: {}'.format(r))

        # Exporting values to excel sheet
        outSheet.write(i + 1, 2, d)
        outSheet.write(i + 1, 3, t)
        outSheet.write(i + 1, 4, p)
        outSheet.write(i + 1, 5, r)

        # Checking if the image meets the critira
        if d > 0.5 and a == 'VERY_UNLIKELY' and j == 'VERY_UNLIKELY' and su == 'VERY_UNLIKELY' and so == 'VERY_UNLIKELY' and h == 'VERY_UNLIKELY' and b == 'VERY_UNLIKELY' and u == 'VERY_UNLIKELY' and  t < 5 and t > -5 and p < 5 and p > -5 and r < 5 and r > -5:
            leftImages.append('1')
        else:
            leftImages.append('0')
        
        # Face bounds are not needed, but just in case ¯\_(ツ)_/¯
        vertices = (['({},{})'.format(vertex.x, vertex.y)
                    for vertex in face.bounding_poly.vertices])
        print('face bounds: {} \n'.format(','.join(vertices)))
        break

    # A flag to mark finding at least one face; because some images are so distorted that no face can be detected
    if flag == 0:
        # Appending zero for distorted image; in order not to mess the list ordering
        leftImages.append('0')
        outSheet.write(i + 1, 2, 0)
        outSheet.write(i + 1, 3, 99)
        outSheet.write(i + 1, 4, 99)
        outSheet.write(i + 1, 5, 99)
        print('Detection confidence: 0')
        print('anger: UNKNOWN')
        print('joy: UNKNOWN')
        print('surprise: UNKNOWN')
        print('sorrow: UNKNOWN')
        print('headwear: UNKNOWN')
        print('blurred: UNKNOWN')
        print('under exposed: UNKNOWN')
        print('tilt angle: 99')
        print('pan angle: 99')
        print('roll angle: 99')
    
    outSheet.write(i + 1, 1, leftImages[i])
    
    # Error Handling
    if response.error.message:
        raise Exception(
            '{}\nFor more info on error messages, check: '
            'https://cloud.google.com/apis/design/errors'.format(
                response.error.message))

import os, sys, xlsxwriter

outWorkbook = xlsxwriter.Workbook('Output.xlsx')
outSheet = outWorkbook.add_worksheet()
# Exporting the path, boolean value, Detection confidence, and angles of each image, to be used for deletion
outSheet.write('A1','Path')
outSheet.write('B1','Bool')
outSheet.write('C1','Detection confidence')
outSheet.write('D1','Tilt')
outSheet.write('E1','Pan')
outSheet.write('F1','Roll')

# Redirecting stdout to output to the text file
sys.stdout = open('Output.txt', 'w')
# Iterator to keep track of images
i = 0
leftImages = []

from openpyxl import load_workbook
# Change to the path of the directory containing the excel file
wb = load_workbook("listing of folder 00000.xlsx")
ws = wb['Sheet1']
columnA = ws['A']
# Getting the names of the images
names = [columnA[x].value for x in range(len(columnA))]

columnB = ws['B']
# Getting the links of the images
images = [columnB[x].value for x in range(len(columnB))]

# Zipping the names and images lists in a dictionary
pair = zip(names,images)
data = dict(pair)

# For each image, calling the API and indexing each image with its number
for name, image in data.items():
    print(name)
    outSheet.write(i + 1, 0, name)
    # An exception handling block; because it keeps getting timeout error
    try:
        detect_faces(image)
    except Exception as m:
        print(leftImages)
        sys.stdout.close()
        outWorkbook.close()
    else:
        i += 1
print(leftImages)
sys.stdout.close()
outWorkbook.close()

'''
# Get Google Drive links (run in Apps Script)

function listFolderContents() {
  var foldername = '02000';
  var folderlisting = 'listing of folder ' + foldername;
  
  var folders = DriveApp.getFoldersByName(foldername)
  if(folders.hasNext())
  {
    var folder = folders.next();
    var contents = folder.getFiles();

    var ss = SpreadsheetApp.create(folderlisting);
    var sheet = ss.getActiveSheet();
    sheet.appendRow(['name','link']);
    
    var file;
    var name;
    var link;
    var row; 
    while(contents.hasNext())
    {
      file = contents.next();
      name = file.getName();
      link = file.getId();
      sheet.appendRow([name, link]);
    }
  }
}

'''
