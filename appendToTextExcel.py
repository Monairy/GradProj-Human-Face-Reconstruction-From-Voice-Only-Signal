import itertools 
det = []
an = []
j = []
su = []
so = []
h = []
bl = []
u = []
t = []
p = []
r = []
leftImages = []
fullPath = [] 

with open('Output.txt', 'r') as searchfile:
    for line in searchfile:
        if '../input/human-faces-dataset/' in line:
            fullPath.append(line.strip())
        elif 'Detection confidence:' in line:
            det.append(float(line[21:].strip()))
        elif 'anger: ' in line:
            an.append(line[7:].strip())
        elif 'joy: ' in line:
            j.append(line[5:].strip())
        elif 'surprise: ' in line:
            su.append(line[9:].strip())
        elif 'sorrow: ' in line:
            so.append(line[8:].strip())
        elif 'headwear: ' in line:
            h.append(line[10:].strip())
        elif 'blurred: ' in line:
            bl.append(line[9:].strip())
        elif 'under exposed: ' in line:
            u.append(line[15:].strip())
        elif 'tilt angle: ' in line:
            t.append(float(line[12:].strip()))
        elif 'pan angle: ' in line:
            p.append(float(line[11:].strip()))
        elif 'roll angle: ' in line:
            r.append(float(line[12:].strip()))

for (a, b, c, d, e, f, g, h, l, m, n) in itertools.zip_longest(det, an, j, su, so, h, bl, u, t, p, r):
    if a > 0.5 and b == 'VERY_UNLIKELY' and c == 'VERY_UNLIKELY' and d == 'VERY_UNLIKELY' and e == 'VERY_UNLIKELY' and f == 'VERY_UNLIKELY' and g == 'VERY_UNLIKELY' and h == 'VERY_UNLIKELY' and  l < 5 and l > -5 and m < 5 and m > -5 and n < 5 and n > -5:
        leftImages.append('1')
    else:
        leftImages.append('0')


from openpyxl import load_workbook

# Change to the path of the directory containing the excel file
wb = load_workbook("Output.xlsx")
ws = wb['Sheet1']
columnA = ws['A']
# Getting the paths of the images
done = [columnA[x].value for x in range(len(columnA))]

# Initialize i with the first empty row
#i = len(done)
if len(fullPath) == len(det) == len(t) == len(p) == len(r):
    length = len(fullPath)
else:
    length = len(fullPath) - 1
    
for i in range(length):
        ws.append([fullPath[i], leftImages[i], det[i], t[i], p[i], r[i]])
        
print(ws.max_row)

wb.save('Output.xlsx')

filename = ['Output.txt']
with open('r1.txt', 'a') as outfile:
    for fname in filename:
        with open(fname) as infile:
            for line in infile:
                outfile.write(line)
outfile.close()
infile.close()

import os
os.remove('Output.txt')
